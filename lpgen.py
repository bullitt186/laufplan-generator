import sys
import csv
import argparse
from datetime import datetime, timedelta
import openpyxl
from icalendar import Calendar, Event
import gettext

# Constants for heart rate zones
HEART_RATE_ZONES = {
    "Zone 1": (50, 60),
    "Zone 2": (61, 70),
    "Zone 3": (71, 80),
    "Zone 4": (81, 90),
    "Zone 5": (91, 100)
}

# Set up internationalization
# I added internationalization using the `gettext` library, which is a standard Python library for handling translations. The `_` function is an alias for `gettext.gettext`, which is used to mark strings for translation. To actually translate the strings, you'll need to create translation files (`.mo` and `.po`) for each language you want to support.

_ = gettext.gettext

def parse_args():
    """
    Parse command-line arguments.

    Returns:
        argparse.Namespace: Parsed arguments.
    """
    parser = argparse.ArgumentParser(
        description=_("Convert a training plan for running from CSV or Excel to an iCal calendar file.")
    )
    parser.add_argument(
        "input_file", help=_("CSV or Excel file containing the training plan"), type=str
    )
    return parser.parse_args()

def read_input_file(input_file):
    """
    Read the input file (CSV or Excel) and return its contents as a list of dictionaries.

    Args:
        input_file (str): Path to the input file.

    Returns:
        list: A list of dictionaries containing the data from the input file.
    """
    if input_file.endswith(".csv"):
        with open(input_file, newline="") as csvfile:
            return list(csv.DictReader(csvfile))
    elif input_file.endswith(".xlsx"):
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
        data = []
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2):
            data.append({headers[i]: cell.value for i, cell in enumerate(row)})
        return data
    else:
        raise ValueError(_("Unsupported file format. Please use .csv or .xlsx files."))

def get_next_monday(date):
    """
    Get the next Monday after the given date.

    Args:
        date (datetime.date): Input date.

    Returns:
        datetime.date: The next Monday after the input date.
    """
    return date + timedelta(days=(7 - date.weekday()))

def get_heart_rate_zone(percentage):
    """
    Get the heart rate zone for the given heart rate percentage.

    Args:
        percentage (float): Heart rate percentage.

    Returns:
        str: Heart rate zone.
    """
    for zone, (min_percent, max_percent) in HEART_RATE_ZONES.items():
        if min_percent <= percentage <= max_percent:
            return zone
    return None


def create_calendar_entry(data, max_hr, start_date, prefix):
    """
    Create a calendar event based on the input data, maximum heart rate, start date, and prefix.

    Args:
        data (dict): Dictionary containing the input data for the event.
        max_hr (int): Maximum heart rate.
        start_date (datetime.date): Start date of the training plan.
        prefix (str): Prefix for the event title.

    Returns:
        Event: An icalendar Event object with the specified data.
    """
    event = Event()
    week = int(data["week"])
    day = int(data["day"])
    date_offset = timedelta(weeks=(week - 1), days=(day - 1))
    event_date = start_date + date_offset
    event.add("dtstart", event_date, parameters={"VALUE": "DATE"})
    event.add("dtend", event_date + timedelta(days=1), parameters={"VALUE": "DATE"})

    if prefix != "":
        event_title = f"{prefix}: {data['training']} / {data['distance']}km"
    else:
        event_title = f"{data['training']} / {data['distance']}km"
    event.add("summary", event_title)

    hr_percentage = float(data["heartrate"])
    hr_bpm = max_hr * hr_percentage / 100
    hr_zone = get_heart_rate_zone(hr_percentage)
    distance = data["distance"]
    details = f"- {data['details']}\n- {_('Heart rate')}: {hr_percentage}% ({hr_bpm} bpm)\n- {hr_zone}\n- {_('Distance')}: {distance} km"
    event.add("description", details)

    return event

def main():
    args = parse_args()
    input_data = read_input_file(args.input_file)
    max_hr = int(input(_("Enter your maximum heart rate (in bpm): ")))
    start_date_str = input(_("Enter the start date (format: yyyy-mm-dd): "))
    start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
    if start_date.weekday() != 0:
        start_date = get_next_monday(start_date)

    calendar = Calendar()
    calendar.add("prodid", "-//Training Plan//example.com//")
    calendar.add("version", "2.0")

    prefix = _("")

    for data in input_data:
        event = create_calendar_entry(data, max_hr, start_date, prefix)
        calendar.add_component(event)

    ical_filename = f"Laufplan_{start_date}.ics"
    with open(ical_filename, "wb") as ical_file:
        ical_file.write(calendar.to_ical())

    print(_("Generated iCal file '{ical_filename}' based on the provided training plan.").format(ical_filename=ical_filename))

if __name__ == "__main__":
    main()
